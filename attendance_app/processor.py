from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from attendance_app.config import AttendanceConfig
from attendance_app.models import (
    DetailRow,
    GenerationResult,
    ParsedHoliday,
    ParsedNonWorkday,
    ParsedWorkbook,
    SummaryStats,
)


STAT_MONTH = date(2026, 6, 1)
TIME_RE = re.compile(r"(?<!\d)([0-2]?\d:[0-5]\d)(?!\d)")


def extract_times(value: object) -> list[time]:
    if value is None:
        return []
    times: list[time] = []
    for raw in TIME_RE.findall(str(value)):
        hour_s, minute_s = raw.split(":")
        hour = int(hour_s)
        minute = int(minute_s)
        if hour < 24:
            times.append(time(hour, minute))
    return times


def expand_cell_timeline(
    source_date: date, times: list[time], config: AttendanceConfig
) -> list[tuple[date, time]]:
    if not times:
        return []

    first_trailing_early_index = len(times)
    while (
        first_trailing_early_index > 0
        and times[first_trailing_early_index - 1] < config.overnight_cutoff
    ):
        first_trailing_early_index -= 1

    has_same_day_punch_before_trailing_early = any(
        punch >= config.overnight_cutoff for punch in times[:first_trailing_early_index]
    )
    if not has_same_day_punch_before_trailing_early:
        first_trailing_early_index = len(times)

    return [
        (
            source_date + timedelta(days=1)
            if index >= first_trailing_early_index
            else source_date,
            punch_time,
        )
        for index, punch_time in enumerate(times)
    ]


def parse_report_start(ws: Any) -> date:
    title = str(ws["A1"].value or "")
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", title)
    if not match:
        return STAT_MONTH
    year, month, day = map(int, match.groups())
    return date(year, month, day)


def parse_report_range(ws: Any) -> tuple[date, date]:
    title = str(ws["A1"].value or "")
    matches = re.findall(r"(\d{4})-(\d{2})-(\d{2})", title)
    if len(matches) >= 2:
        start_year, start_month, start_day = map(int, matches[0])
        end_year, end_month, end_day = map(int, matches[1])
        return (
            date(start_year, start_month, start_day),
            date(end_year, end_month, end_day),
        )
    start = parse_report_start(ws)
    return start, start


def parse_generated_at_date(ws: Any) -> date | None:
    text = str(ws["A2"].value or "")
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if not match:
        return None
    year, month, day = map(int, match.groups())
    return date(year, month, day)


def is_holiday_header(header: object) -> bool:
    text = str(header or "").strip()
    return bool(text and not text.isdigit() and text not in {"六", "日"})


def get_day_type(current_date: date, header: object) -> str:
    text = str(header or "").strip()
    if is_holiday_header(header):
        return text
    if current_date.weekday() == 5:
        return "周六"
    if current_date.weekday() == 6:
        return "周日"
    return "工作日"


def parse_workbook(input_file: Path) -> ParsedWorkbook:
    source = load_workbook(input_file, data_only=True, read_only=True)
    ws = source.active
    report_start, report_end = parse_report_range(ws)

    holidays: list[ParsedHoliday] = []
    weekend_dates: list[date] = []
    non_workdays: list[ParsedNonWorkday] = []
    date_count = max(0, ws.max_column - 6)
    for offset, col in enumerate(range(7, ws.max_column + 1)):
        current_date = report_start + timedelta(days=offset)
        header = ws.cell(row=4, column=col).value
        day_type = get_day_type(current_date, header)
        if current_date.weekday() >= 5:
            weekend_dates.append(current_date)
        if is_holiday_header(header):
            holidays.append(ParsedHoliday(date=current_date, label=str(header).strip()))
        if day_type != "工作日":
            non_workdays.append(ParsedNonWorkday(date=current_date, label=day_type))

    employee_count = 0
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            employee_count += 1

    source.close()
    return ParsedWorkbook(
        report_start=report_start,
        report_end=report_end,
        suggested_start_date=report_start,
        suggested_end_date=report_end,
        suggested_ignore_dates=[],
        holidays=holidays,
        weekend_dates=weekend_dates,
        non_workdays=non_workdays,
        employee_count=employee_count,
        date_count=date_count,
    )


def fmt_date(d: date) -> str:
    return f"{d.month}月{d.day}日"


def fmt_dt_for_shift(dt: datetime, base_date: date) -> str:
    if dt.date() > base_date:
        return f"次日 {dt:%H:%M}"
    return dt.strftime("%H:%M")


def workday_overtime(
    punches: list[datetime], base_date: date, config: AttendanceConfig
) -> tuple[float, bool, str]:
    same_day_punches = [dt for dt in punches if dt.date() == base_date]
    next_day_early_punches = [
        dt
        for dt in punches
        if dt.date() == base_date + timedelta(days=1)
        and dt.time() < config.overnight_cutoff
    ]

    if next_day_early_punches:
        return 4.0, True, "次日凌晨打卡，工作日加班按4小时折算"

    if not same_day_punches:
        return 0.0, False, ""

    last_same_day = max(same_day_punches)
    if last_same_day.time() < config.overtime_start_time:
        return 0.0, False, ""
    if last_same_day.time() >= config.workday_meal_after:
        return 2.0, True, ""
    if last_same_day.time() >= config.workday_overtime_2h_after:
        return 2.0, False, ""
    return 0.0, False, ""


def restday_overtime(last_dt: datetime, base_date: date) -> tuple[float, bool]:
    noon = datetime.combine(base_date, time(12, 0))
    five_pm = datetime.combine(base_date, time(17, 0))
    nine_pm = datetime.combine(base_date, time(21, 0))
    if last_dt >= nine_pm:
        return 8.0, True
    if last_dt >= five_pm:
        return 6.0, False
    if last_dt >= noon:
        return 3.0, False
    return 0.0, False


def is_late_for_workday(
    punches: list[datetime], base_date: date, config: AttendanceConfig
) -> bool:
    same_day_punches = [dt for dt in punches if dt.date() == base_date]
    if not same_day_punches:
        return False
    return min(same_day_punches).time() > config.work_start_time


def has_normal_start_punch(punches: list[time], config: AttendanceConfig) -> bool:
    return any(config.overnight_cutoff <= punch < config.work_start_time for punch in punches)


def has_non_early_punch(punches: list[time], config: AttendanceConfig) -> bool:
    return any(punch >= config.overnight_cutoff for punch in punches)


def previous_day_is_incomplete(punches: list[time], config: AttendanceConfig) -> bool:
    return len(punches) == 1


def assign_punch_base_date(
    source_date: date,
    punch_time: time,
    natural_day_times: dict[date, list[time]],
    config: AttendanceConfig,
) -> date:
    if punch_time >= config.overnight_cutoff:
        return source_date

    today_times = natural_day_times.get(source_date, [])
    if has_normal_start_punch(today_times, config):
        return source_date - timedelta(days=1)

    if has_non_early_punch(today_times, config):
        return source_date

    previous_times = natural_day_times.get(source_date - timedelta(days=1), [])
    if previous_day_is_incomplete(previous_times, config):
        return source_date - timedelta(days=1)
    return source_date


def style_sheet(ws: Any, freeze_cell: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(45, len(value)))
        ws.column_dimensions[col_letter].width = max(10, min(45, max_len + 2))


def generate_summary(
    input_file: Path, output_dir: Path, config: AttendanceConfig | None = None
) -> GenerationResult:
    active_config = config or AttendanceConfig()
    source = load_workbook(input_file, data_only=True)
    ws = source.active
    report_start = parse_report_start(ws)

    source_date_columns: list[tuple[int, date, object]] = []
    date_headers: dict[date, object] = {}
    for offset, col in enumerate(range(7, ws.max_column + 1)):
        current_date = report_start + timedelta(days=offset)
        header = ws.cell(row=4, column=col).value
        source_date_columns.append((col, current_date, header))
        date_headers[current_date] = header

    ignored_dates = set(active_config.ignore_dates)
    report_dates = []
    for offset in range(ws.max_column - 6):
        current = report_start + timedelta(days=offset)
        if current in ignored_dates:
            continue
        if active_config.start_date is not None and current < active_config.start_date:
            continue
        if active_config.end_date is not None and current > active_config.end_date:
            continue
        report_dates.append(current)

    detail_rows: list[DetailRow] = []
    person_summary: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in range(5, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        department = str(ws.cell(row=row, column=3).value or "").strip()
        employee_id = str(ws.cell(row=row, column=4).value or "").strip()
        key = (name, department, employee_id)
        person_summary.setdefault(
            key,
            {
                "ot_dates": [],
                "absence_dates": [],
                "late_dates": [],
                "meal_dates": [],
                "total_hours": 0.0,
                "meal_count": 0,
            },
        )

        natural_day_times: dict[date, list[time]] = defaultdict(list)
        raw_cells: list[tuple[date, list[tuple[date, time]]]] = []
        for col, source_date, _header in source_date_columns:
            raw_value = ws.cell(row=row, column=col).value
            times = extract_times(raw_value)
            if not times:
                continue
            expanded_times = expand_cell_timeline(source_date, times, active_config)
            for actual_date, punch_time in expanded_times:
                natural_day_times[actual_date].append(punch_time)
            raw_cells.append((source_date, expanded_times))

        assigned: dict[date, dict[str, list[tuple[datetime, str]]]] = defaultdict(
            lambda: {"punches": []}
        )
        for _source_date, expanded_times in raw_cells:
            for actual_date, punch_time in expanded_times:
                base_date = assign_punch_base_date(
                    actual_date, punch_time, natural_day_times, active_config
                )
                punch_dt = datetime.combine(actual_date, punch_time)
                if base_date in report_dates:
                    assigned[base_date]["punches"].append(
                        (punch_dt, f"{fmt_date(actual_date)} {punch_time:%H:%M}")
                    )

        for current_date in report_dates:
            day_data = assigned.get(current_date)
            if not day_data:
                continue
            punch_items = sorted(day_data["punches"], key=lambda item: item[0])
            punches = [item[0] for item in punch_items]
            raw_parts = [item[1] for item in punch_items]
            header = date_headers.get(current_date)
            holiday = is_holiday_header(header)
            rest_day = current_date.weekday() >= 5 or holiday
            day_type = get_day_type(current_date, header)
            absence = False
            overtime_hours = 0.0
            meal = False
            note = ""

            last_dt = punches[-1]
            last_punch = fmt_dt_for_shift(last_dt, current_date)
            late = False if rest_day else is_late_for_workday(punches, current_date, active_config)

            if len(punches) == 1 and not rest_day:
                absence = True
                note = "工作日只有一次打卡，按旷工标记"
            else:
                if rest_day:
                    overtime_hours, meal = restday_overtime(last_dt, current_date)
                else:
                    overtime_hours, meal, note = workday_overtime(
                        punches, current_date, active_config
                    )
                if last_dt.date() > current_date:
                    note = note or "次日凌晨打卡，按跨天记录处理"

            if overtime_hours > 0 or meal or absence or late:
                detail_rows.append(
                    DetailRow(
                        name=name,
                        department=department,
                        employee_id=employee_id,
                        work_date=current_date,
                        day_type=day_type,
                        raw_punches="\n".join(raw_parts),
                        last_punch=last_punch,
                        overtime_hours=overtime_hours,
                        meal=meal,
                        meal_amount=active_config.meal_allowance_amount if meal else 0,
                        absent=absence,
                        late=late,
                        note=note,
                    )
                )

            summary = person_summary[key]
            if overtime_hours > 0:
                summary["ot_dates"].append(f"{fmt_date(current_date)}加班{overtime_hours:g}小时")
                summary["total_hours"] = float(summary["total_hours"]) + overtime_hours
            if absence:
                summary["absence_dates"].append(f"{fmt_date(current_date)}旷工")
            if late:
                summary["late_dates"].append(f"{fmt_date(current_date)}迟到")
            if meal:
                summary["meal_dates"].append(f"{fmt_date(current_date)}有餐补")
                summary["meal_count"] = int(summary["meal_count"]) + 1

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / active_config.output_filename
    write_workbook(output_path, detail_rows, person_summary, active_config)
    stats = build_stats(detail_rows, len(person_summary))
    return GenerationResult(output_path=output_path, detail_rows=detail_rows, stats=stats)


def build_stats(detail_rows: list[DetailRow], people: int) -> SummaryStats:
    meal_records = sum(1 for row in detail_rows if row.meal)
    return SummaryStats(
        people=people,
        detail_records=len(detail_rows),
        overtime_records=sum(1 for row in detail_rows if row.overtime_hours > 0),
        absence_records=sum(1 for row in detail_rows if row.absent),
        late_records=sum(1 for row in detail_rows if row.late),
        meal_records=meal_records,
        total_overtime_hours=round(sum(row.overtime_hours for row in detail_rows), 2),
        total_meal_amount=sum(row.meal_amount for row in detail_rows),
    )


def write_workbook(
    output_path: Path,
    detail_rows: list[DetailRow],
    person_summary: dict[tuple[str, str, str], dict[str, Any]],
    config: AttendanceConfig,
) -> None:
    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = "加班明细"
    detail_ws.append(
        [
            "姓名",
            "部门",
            "工号",
            "日期",
            "日期类型",
            "原始打卡时间",
            "下班最后打卡时间",
            "折算加班时长",
            "是否餐补",
            "餐补金额",
            "是否旷工",
            "是否迟到",
            "备注",
        ]
    )
    for row in detail_rows:
        detail_ws.append(
            [
                row.name,
                row.department,
                row.employee_id,
                row.work_date,
                row.day_type,
                row.raw_punches,
                row.last_punch,
                row.overtime_hours,
                "是" if row.meal else "否",
                row.meal_amount,
                "是" if row.absent else "否",
                "是" if row.late else "否",
                row.note,
            ]
        )
    for cell in detail_ws["D"][1:]:
        cell.number_format = "m月d日"
    for cell in detail_ws["H"][1:]:
        cell.number_format = "0.00"
    for cell in detail_ws["J"][1:]:
        cell.number_format = "#,##0"
    style_sheet(detail_ws, "A2")

    summary_ws = wb.create_sheet("人员汇总")
    summary_ws.append(
        [
            "姓名",
            "部门",
            "工号",
            "加班日期",
            "旷工日期",
            "迟到日期",
            "餐补日期",
            "本月加班时长合计",
            "餐补次数",
            "餐补金额合计",
        ]
    )
    for (name, department, employee_id), summary in sorted(person_summary.items()):
        total_hours = round(float(summary["total_hours"]), 2)
        meal_count = int(summary["meal_count"])
        if total_hours == 0 and meal_count == 0 and not summary["absence_dates"] and not summary["late_dates"]:
            continue
        summary_ws.append(
            [
                name,
                department,
                employee_id,
                "\n".join(summary["ot_dates"]),
                "\n".join(summary["absence_dates"]),
                "\n".join(summary["late_dates"]),
                "\n".join(summary["meal_dates"]),
                total_hours,
                meal_count,
                meal_count * config.meal_allowance_amount,
            ]
        )
    for cell in summary_ws["H"][1:]:
        cell.number_format = "0.00"
    for cell in summary_ws["I"][1:]:
        cell.number_format = "#,##0"
    for cell in summary_ws["J"][1:]:
        cell.number_format = "#,##0"
    style_sheet(summary_ws, "A2")

    rules_ws = wb.create_sheet("统计口径")
    rules = [
        ["项目", "口径"],
        ["统计范围", "默认忽略配置中的日期；样例默认忽略2026年6月29日"],
        ["工作日", "按真实日期为周一至周五，且表头不是节假日文本"],
        ["休息日/节假日", "按真实日期为周六、周日，或表头为非数字且不是“六/日”的节假日文本"],
        ["工作日加班", "达到加班起始时间后才进入加班判断；达到2小时加班阈值算2小时；次日凌晨打卡截止前有打卡记录算4小时"],
        ["工作日餐补", "最后打卡时间达到工作日餐补阈值或更晚，或次日凌晨打卡截止前有打卡记录，享受餐补"],
        ["休息日/节假日加班", "最后打卡到12:00算3小时；17:00以后算6小时；21:00以后算8小时"],
        ["休息日/节假日餐补", "最后打卡达到21:00或更晚，享受餐补"],
        ["迟到", "工作日当天第一条凌晨打卡截止及之后的打卡晚于上班时间，标记迟到；周末/节假日不判迟到"],
        ["凌晨打卡截止", "次日凌晨打卡截止前打卡是跨天加班的打卡约束"],
        ["旷工", "工作日只有一个打卡时间视为旷工；周末/节假日只有一个打卡时间不视为旷工"],
    ]
    for rule_row in rules:
        rules_ws.append(rule_row)
    style_sheet(rules_ws, "A2")

    wb.save(output_path)
