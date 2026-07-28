from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LEAVE_TYPES = (
    "事假",
    "调休",
    "病假",
    "年假",
    "产假",
    "陪产假",
    "婚假",
    "例假",
    "丧假",
    "哺乳假",
    "高温假",
)
LEAVE_PATTERN = re.compile(
    rf"({'|'.join(LEAVE_TYPES)})(\d{{2}})-(\d{{2}})\s+(\d{{2}}:\d{{2}})"
    rf"到(\d{{2}})-(\d{{2}})\s+(\d{{2}}:\d{{2}})"
)
PUNCH_RESULT_PATTERN = re.compile(r"\(([^()]*)\)\s*$")
WEEKEND_TYPES = {"周六", "周日", "周末"}
ANOMALY_STATUSES = {"迟到", "缺卡", "早退", "旷工", "异常"}


@dataclass
class MonthlyEmployee:
    name: str
    department: str
    employee_id: str
    user_id: str
    has_monthly_source: bool = True
    leave_totals: dict[str, float] = field(default_factory=dict)
    absence_days: float = 0.0
    daily_results: dict[date, str] = field(default_factory=dict)


@dataclass
class MonthlyWorkbook:
    report_start: date
    report_end: date
    employees: list[MonthlyEmployee]
    company_name: str


def parse_report_range_text(value: object) -> tuple[date, date]:
    matches = re.findall(r"(\d{4})-(\d{2})-(\d{2})", str(value or ""))
    if len(matches) < 2:
        raise ValueError("月度汇总表标题中未找到完整统计日期范围")
    start_year, start_month, start_day = map(int, matches[0])
    end_year, end_month, end_day = map(int, matches[1])
    report_start = date(start_year, start_month, start_day)
    report_end = date(end_year, end_month, end_day)
    if report_end < report_start:
        raise ValueError("月度汇总表标题中的统计结束日期早于开始日期")
    return report_start, report_end


def parse_company_name(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"^减离职人员\s*", "", stem)
    stem = re.sub(r"_月度汇总_.*$", "", stem)
    return stem.strip() or "考勤汇总"


def numeric(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0


def parse_monthly_workbook(input_file: Path) -> MonthlyWorkbook:
    source = load_workbook(input_file, data_only=True)
    ws = source.active
    try:
        report_start, report_end = parse_report_range_text(ws["A1"].value)
    except Exception:
        source.close()
        raise

    daily_start_col: int | None = None
    leave_columns: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        group_header = str(ws.cell(row=3, column=col).value or "").strip()
        sub_header = str(ws.cell(row=4, column=col).value or "").strip()
        if group_header == "考勤结果":
            daily_start_col = col
        if sub_header:
            for leave_type in LEAVE_TYPES:
                if sub_header.startswith(f"{leave_type}("):
                    leave_columns[leave_type] = col

    if daily_start_col is None:
        source.close()
        raise ValueError("月度汇总表中未找到“考勤结果”每日明细")

    absence_col = next(
        (
            col
            for col in range(1, ws.max_column + 1)
            if str(ws.cell(row=3, column=col).value or "").strip() == "旷工天数"
        ),
        None,
    )
    expected_days = (report_end - report_start).days + 1
    if ws.max_column - daily_start_col + 1 < expected_days:
        source.close()
        raise ValueError("月度汇总表的每日考勤结果列少于统计日期天数")

    employees: list[MonthlyEmployee] = []
    for row in range(5, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=1).value
        if not raw_name:
            continue
        daily_results: dict[date, str] = {}
        for offset in range(expected_days):
            current_date = report_start + timedelta(days=offset)
            value = ws.cell(row=row, column=daily_start_col + offset).value
            if value not in (None, ""):
                daily_results[current_date] = str(value).strip()
        employees.append(
            MonthlyEmployee(
                name=str(raw_name).strip(),
                department=str(ws.cell(row=row, column=3).value or "").strip(),
                employee_id=str(ws.cell(row=row, column=4).value or "").strip(),
                user_id=str(ws.cell(row=row, column=6).value or "").strip(),
                leave_totals={
                    leave_type: numeric(ws.cell(row=row, column=col).value)
                    for leave_type, col in leave_columns.items()
                },
                absence_days=numeric(
                    ws.cell(row=row, column=absence_col).value if absence_col else None
                ),
                daily_results=daily_results,
            )
        )

    source.close()
    return MonthlyWorkbook(
        report_start=report_start,
        report_end=report_end,
        employees=employees,
        company_name=parse_company_name(input_file),
    )


def resolve_month_day(
    month: int, day: int, reference_date: date
) -> date:
    year = reference_date.year
    if month - reference_date.month > 6:
        year -= 1
    elif reference_date.month - month > 6:
        year += 1
    return date(year, month, day)


def parse_leave_intervals(raw_result: str, reference_date: date) -> list[tuple[str, datetime, datetime]]:
    intervals: list[tuple[str, datetime, datetime]] = []
    for match in LEAVE_PATTERN.finditer(raw_result):
        (
            leave_type,
            start_month,
            start_day,
            start_time,
            end_month,
            end_day,
            end_time,
        ) = match.groups()
        start_date = resolve_month_day(int(start_month), int(start_day), reference_date)
        end_date = resolve_month_day(int(end_month), int(end_day), reference_date)
        intervals.append(
            (
                leave_type,
                datetime.combine(start_date, time.fromisoformat(start_time)),
                datetime.combine(end_date, time.fromisoformat(end_time)),
            )
        )
    return intervals


def overlaps(start: datetime, end: datetime, period_start: datetime, period_end: datetime) -> bool:
    return start < period_end and end > period_start


def leave_status_for_period(
    intervals: list[tuple[str, datetime, datetime]],
    period_start: datetime,
    period_end: datetime,
) -> str | None:
    matching = {
        leave_type
        for leave_type, start, end in intervals
        if overlaps(start, end, period_start, period_end)
    }
    if len(matching) > 1:
        return "异常"
    return next(iter(matching), None)


def parse_result_punches(raw_result: str) -> tuple[bool, bool]:
    match = PUNCH_RESULT_PATTERN.search(raw_result)
    if not match:
        return False, False
    parts = [part.strip() for part in match.group(1).split(",")]
    first_present = bool(parts and parts[0] not in {"", "-"})
    last_present = bool(len(parts) > 1 and parts[-1] not in {"", "-"})
    if len(parts) == 1:
        last_present = first_present
    return first_present, last_present


def derive_half_day_statuses(
    raw_result: str,
    current_date: date,
    day_type: str,
) -> tuple[str, str]:
    if day_type != "工作日":
        return "", ""

    intervals = parse_leave_intervals(raw_result, current_date)
    morning_leave = leave_status_for_period(
        intervals,
        datetime.combine(current_date, time(8, 30)),
        datetime.combine(current_date, time(12, 0)),
    )
    afternoon_leave = leave_status_for_period(
        intervals,
        datetime.combine(current_date, time(13, 30)),
        datetime.combine(current_date, time(18, 0)),
    )

    first_present, last_present = parse_result_punches(raw_result)
    if "旷工" in raw_result and "旷工迟到" not in raw_result:
        morning = "旷工"
        afternoon = "旷工"
    elif any(status in raw_result for status in ("出差", "外出", "外勤")):
        morning = "√"
        afternoon = "√"
    elif "正常" in raw_result:
        morning = "√"
        afternoon = "√"
    else:
        morning = "√" if first_present else "异常"
        afternoon = "√" if last_present else "异常"

    if "上班缺卡" in raw_result:
        morning = "缺卡"
    elif "上班迟到" in raw_result or "严重迟到" in raw_result or "旷工迟到" in raw_result:
        morning = "迟到"
    if "下班缺卡" in raw_result:
        afternoon = "缺卡"
    elif "早退" in raw_result:
        afternoon = "早退"

    return morning_leave or morning, afternoon_leave or afternoon


def employee_leave_days(employee: MonthlyEmployee) -> dict[str, float]:
    hourly_types = {"事假", "调休", "病假", "哺乳假"}
    return {
        leave_type: round(value / 8 if leave_type in hourly_types else value, 2)
        for leave_type, value in employee.leave_totals.items()
    }


def value_or_blank(value: float) -> float | str:
    rounded = round(value, 2)
    return rounded if rounded else ""


def write_attendance_summary_sheet(
    wb: Workbook,
    monthly: MonthlyWorkbook,
    display_dates: list[date],
    date_headers: dict[date, object],
    computed_by_user_id: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("汇总表", 0)
    ws.sheet_view.showGridLines = False

    date_start_col = 15
    date_end_col = date_start_col + len(display_dates) - 1
    spacer_col = date_end_col + 1
    remarks_col = spacer_col + 1
    signature_col = spacer_col + 2
    last_col_letter = get_column_letter(signature_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=signature_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=signature_col)
    ws.cell(row=1, column=1, value=f"{display_dates[0].year}年{display_dates[0].month}月考勤汇总表")
    ws.cell(row=2, column=1, value=monthly.company_name)
    ws.cell(row=1, column=1).font = Font(name="宋体", size=22, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).font = Font(name="宋体", size=14, bold=True)

    fixed_headers = [
        "序号",
        "部门",
        "姓名",
        "法假\n加班\n（天）",
        "平时\n加班\n（小时）",
        "周末加班\n（小时）",
        "补休\n（天）",
        "年假\n（天）",
        "其他假\n（天）",
        "事假\n（天）",
        "病假\n（天）",
        "产假\n（天）",
        "旷工\n（天）",
    ]
    for col, header in enumerate(fixed_headers, start=1):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        ws.cell(row=3, column=col, value=header)
    ws.cell(row=3, column=14, value="星期")
    ws.cell(row=4, column=14, value="日期")
    ws.merge_cells(start_row=3, start_column=remarks_col, end_row=4, end_column=remarks_col)
    ws.merge_cells(start_row=3, start_column=signature_col, end_row=4, end_column=signature_col)
    ws.cell(row=3, column=remarks_col, value="备注（异常/补贴）")
    ws.cell(row=3, column=signature_col, value="员工签名")

    weekday_labels = "一二三四五六日"
    yellow_fill = PatternFill("solid", fgColor="FFF200")
    for col, current_date in enumerate(display_dates, start=date_start_col):
        ws.cell(row=3, column=col, value=weekday_labels[current_date.weekday()])
        ws.cell(row=4, column=col, value=current_date.day)
        day_type = day_type_for_header(date_headers.get(current_date))
        if day_type != "工作日":
            for row in range(3, 5 + len(monthly.employees) * 2):
                ws.cell(row=row, column=col).fill = yellow_fill

    normal_font = Font(name="宋体", size=9)
    anomaly_font = Font(name="宋体", size=9, bold=True, color="C00000")
    leave_font = Font(name="宋体", size=9, color="1F4E78")
    status_fill = PatternFill("solid", fgColor="FCE4D6")

    for index, employee in enumerate(monthly.employees, start=1):
        morning_row = 5 + (index - 1) * 2
        afternoon_row = morning_row + 1
        for col in range(1, 14):
            ws.merge_cells(
                start_row=morning_row,
                start_column=col,
                end_row=afternoon_row,
                end_column=col,
            )
        ws.merge_cells(
            start_row=morning_row,
            start_column=remarks_col,
            end_row=afternoon_row,
            end_column=remarks_col,
        )
        ws.merge_cells(
            start_row=morning_row,
            start_column=signature_col,
            end_row=afternoon_row,
            end_column=signature_col,
        )

        computed = computed_by_user_id.get(employee.user_id, {})
        leave_days = employee_leave_days(employee)
        other_leave_days = sum(
            leave_days.get(leave_type, 0.0)
            for leave_type in ("陪产假", "婚假", "例假", "丧假", "哺乳假", "高温假")
        )
        values: list[object] = [
            index,
            employee.department,
            employee.name,
            value_or_blank(float(computed.get("holiday_days", 0.0))),
            value_or_blank(float(computed.get("workday_hours", 0.0))),
            value_or_blank(float(computed.get("weekend_hours", 0.0))),
            value_or_blank(leave_days.get("调休", 0.0)),
            value_or_blank(leave_days.get("年假", 0.0)),
            value_or_blank(other_leave_days),
            value_or_blank(leave_days.get("事假", 0.0)),
            value_or_blank(leave_days.get("病假", 0.0)),
            value_or_blank(leave_days.get("产假", 0.0)),
            value_or_blank(employee.absence_days),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=morning_row, column=col, value=value)
        ws.cell(row=morning_row, column=14, value="上午")
        ws.cell(row=afternoon_row, column=14, value="下午")

        anomaly_notes: list[str] = []
        for col, current_date in enumerate(display_dates, start=date_start_col):
            day_type = day_type_for_header(date_headers.get(current_date))
            if employee.has_monthly_source:
                raw_result = employee.daily_results.get(current_date, "")
                morning, afternoon = derive_half_day_statuses(
                    raw_result, current_date, day_type
                )
            else:
                morning, afternoon = "", ""
            for row, status, period in (
                (morning_row, morning, "上午"),
                (afternoon_row, afternoon, "下午"),
            ):
                cell = ws.cell(row=row, column=col, value=status)
                cell.font = normal_font
                if status in ANOMALY_STATUSES:
                    cell.font = anomaly_font
                    cell.fill = status_fill
                    anomaly_notes.append(f"{current_date.month}/{current_date.day}{period}{status}")
                elif status in LEAVE_TYPES:
                    cell.font = leave_font
        meal_count = int(computed.get("meal_count", 0))
        meal_amount = int(computed.get("meal_amount", 0))
        remarks: list[str] = []
        if meal_count:
            remarks.append(f"餐补{meal_count}次/{meal_amount}元")
        remarks.extend(anomaly_notes)
        ws.cell(row=morning_row, column=remarks_col, value="；".join(remarks))

    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    last_data_row = 4 + len(monthly.employees) * 2
    for row in ws.iter_rows(min_row=3, max_row=last_data_row, min_col=1, max_col=signature_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row in {3, 4}:
                cell.font = Font(name="宋体", size=9, bold=True)
    for row in range(5, last_data_row + 1):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 23
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 22

    widths = {
        1: 6,
        2: 18,
        3: 10,
        4: 8,
        5: 8,
        6: 9,
        7: 7,
        8: 7,
        9: 7,
        10: 7,
        11: 7,
        12: 7,
        13: 7,
        14: 7,
        remarks_col: 24,
        signature_col: 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(date_start_col, date_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.column_dimensions[get_column_letter(spacer_col)].width = 2

    for row in range(5, last_data_row + 1, 2):
        for col in range(4, 14):
            ws.cell(row=row, column=col).number_format = "0.##"
    ws.freeze_panes = "O5"
    ws.auto_filter.ref = f"A4:{last_col_letter}{last_data_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"


def day_type_for_header(header: object) -> str:
    text = str(header or "").strip()
    if not text or text.isdigit():
        return "工作日"
    if text in {"六", "周六"}:
        return "周六"
    if text in {"日", "周日"}:
        return "周日"
    if text == "周末":
        return "周末"
    return text
