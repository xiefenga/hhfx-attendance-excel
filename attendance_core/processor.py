from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import replace
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from attendance_core.config import AttendanceConfig
from attendance_core.models import (
    DetailRow,
    GenerationResult,
    ParsedHoliday,
    ParsedNonWorkday,
    ParsedWorkbook,
    SummaryStats,
)
from attendance_core.monthly import (
    MonthlyWorkbook,
    parse_monthly_workbook,
    write_attendance_summary_sheet,
)


STAT_MONTH = date(2026, 6, 1)
WEEKEND_HEADER_LABELS = {
    "六": "周六",
    "日": "周日",
    "周六": "周六",
    "周日": "周日",
    "周末": "周末",
}
WEEKEND_DAY_TYPES = set(WEEKEND_HEADER_LABELS.values())


def parse_time_at(
    text: str, index: int, *, allow_leading_digit: bool = False
) -> tuple[time, int] | None:
    if index > 0 and text[index - 1].isdigit() and not allow_leading_digit:
        return None

    for hour_length in (2, 1):
        colon_index = index + hour_length
        minute_start = colon_index + 1
        minute_end = minute_start + 2
        if minute_end > len(text):
            continue
        if text[colon_index] != ":":
            continue
        hour_text = text[index:colon_index]
        minute_text = text[minute_start:minute_end]
        if not hour_text.isdigit() or not minute_text.isdigit():
            continue
        hour = int(hour_text)
        minute = int(minute_text)
        if hour < 24 and minute < 60:
            return time(hour, minute), minute_end
    return None


def extract_times(value: object) -> list[time]:
    if value is None:
        return []
    text = str(value)
    times: list[time] = []
    index = 0
    last_match_end = -1
    while index < len(text):
        candidate = parse_time_at(
            text,
            index,
            allow_leading_digit=index == last_match_end,
        )
        if candidate is None:
            index += 1
            continue
        punch_time, end = candidate
        times.append(punch_time)
        index = end
        last_match_end = end
    return times


def expand_cell_timeline(
    source_date: date, times: list[time], config: AttendanceConfig
) -> list[tuple[date, time]]:
    if not times:
        return []

    first_trailing_early_index = len(times)
    while (
        first_trailing_early_index > 0
        and times[first_trailing_early_index - 1] < config.overnight_cutoff
    ):
        first_trailing_early_index -= 1

    has_same_day_punch_before_trailing_early = any(
        punch >= config.overnight_cutoff for punch in times[:first_trailing_early_index]
    )
    if not has_same_day_punch_before_trailing_early:
        first_trailing_early_index = len(times)

    return [
        (
            source_date + timedelta(days=1)
            if index >= first_trailing_early_index
            else source_date,
            punch_time,
        )
        for index, punch_time in enumerate(times)
    ]


def parse_report_start(ws: Any) -> date:
    title = str(ws["A1"].value or "")
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", title)
    if not match:
        return STAT_MONTH
    year, month, day = map(int, match.groups())
    return date(year, month, day)


def parse_report_range(ws: Any) -> tuple[date, date]:
    title = str(ws["A1"].value or "")
    matches = re.findall(r"(\d{4})-(\d{2})-(\d{2})", title)
    if len(matches) >= 2:
        start_year, start_month, start_day = map(int, matches[0])
        end_year, end_month, end_day = map(int, matches[1])
        return (
            date(start_year, start_month, start_day),
            date(end_year, end_month, end_day),
        )
    start = parse_report_start(ws)
    return start, start


def parse_generated_at_date(ws: Any) -> date | None:
    text = str(ws["A2"].value or "")
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if not match:
        return None
    year, month, day = map(int, match.groups())
    return date(year, month, day)


def is_holiday_header(header: object) -> bool:
    text = str(header or "").strip()
    return bool(text and not text.isdigit() and text not in WEEKEND_HEADER_LABELS)


def get_day_type(current_date: date, header: object) -> str:
    text = str(header or "").strip()
    if not text or text.isdigit():
        return "工作日"
    if text in WEEKEND_HEADER_LABELS:
        return WEEKEND_HEADER_LABELS[text]
    if is_holiday_header(header):
        return text
    return "工作日"


def parse_workbook(input_file: Path) -> ParsedWorkbook:
    source = load_workbook(input_file, data_only=True)
    ws = source.active
    report_start, report_end = parse_report_range(ws)
    expected_date_count = (report_end - report_start).days + 1
    if ws.max_column - 6 < expected_date_count:
        source.close()
        raise ValueError("打卡时间表的每日打卡列少于统计日期天数")

    holidays: list[ParsedHoliday] = []
    weekend_dates: list[date] = []
    non_workdays: list[ParsedNonWorkday] = []
    date_count = expected_date_count
    for offset, col in enumerate(range(7, 7 + expected_date_count)):
        current_date = report_start + timedelta(days=offset)
        header = ws.cell(row=4, column=col).value
        day_type = get_day_type(current_date, header)
        if day_type in WEEKEND_DAY_TYPES:
            weekend_dates.append(current_date)
        if is_holiday_header(header):
            holidays.append(ParsedHoliday(date=current_date, label=str(header).strip()))
        if day_type != "工作日":
            non_workdays.append(ParsedNonWorkday(date=current_date, label=day_type))

    employee_count = 0
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            employee_count += 1

    source.close()
    return ParsedWorkbook(
        report_start=report_start,
        report_end=report_end,
        suggested_start_date=report_start,
        suggested_end_date=report_end,
        suggested_ignore_dates=[],
        holidays=holidays,
        weekend_dates=weekend_dates,
        non_workdays=non_workdays,
        employee_count=employee_count,
        date_count=date_count,
    )


def workbook_employees(input_file: Path) -> dict[str, str]:
    source = load_workbook(input_file, data_only=True)
    ws = source.active
    employees: dict[str, str] = {}
    for row in range(5, ws.max_row + 1):
        name = str(ws.cell(row=row, column=1).value or "").strip()
        user_id = str(ws.cell(row=row, column=6).value or "").strip()
        if not name:
            continue
        if not user_id:
            source.close()
            raise ValueError(f"打卡时间表中员工“{name}”缺少钉钉 UserId")
        if user_id in employees:
            source.close()
            raise ValueError(f"打卡时间表中存在重复钉钉 UserId：{user_id}")
        employees[user_id] = name
    source.close()
    return employees


def parse_sources(input_file: Path, monthly_file: Path) -> ParsedWorkbook:
    parsed = parse_workbook(input_file)
    monthly = parse_monthly_workbook(monthly_file)
    if (parsed.report_start, parsed.report_end) != (
        monthly.report_start,
        monthly.report_end,
    ):
        raise ValueError("两张钉钉导出表的统计日期范围不一致")

    punch_employees = workbook_employees(input_file)
    monthly_employees: dict[str, str] = {}
    for employee in monthly.employees:
        if not employee.user_id:
            raise ValueError(f"月度汇总表中员工“{employee.name}”缺少钉钉 UserId")
        if employee.user_id in monthly_employees:
            raise ValueError(f"月度汇总表中存在重复钉钉 UserId：{employee.user_id}")
        monthly_employees[employee.user_id] = employee.name

    punch_ids = set(punch_employees)
    monthly_ids = set(monthly_employees)
    punch_only = sorted(punch_employees[user_id] for user_id in punch_ids - monthly_ids)
    monthly_only = sorted(monthly_employees[user_id] for user_id in monthly_ids - punch_ids)
    warnings: list[str] = []
    if punch_only:
        warnings.append(f"仅打卡时间表存在：{'、'.join(punch_only)}")
    if monthly_only:
        warnings.append(f"仅月度汇总表存在：{'、'.join(monthly_only)}")

    return replace(
        parsed,
        monthly_employee_count=len(monthly.employees),
        matched_employee_count=len(punch_ids & monthly_ids),
        punch_only_employees=punch_only,
        monthly_only_employees=monthly_only,
        source_warnings=warnings,
    )


def fmt_date(d: date) -> str:
    return f"{d.month}月{d.day}日"


def fmt_dt_for_shift(dt: datetime, base_date: date) -> str:
    if dt.date() > base_date:
        return f"次日 {dt:%H:%M}"
    return dt.strftime("%H:%M")


def workday_overtime(
    punches: list[datetime], base_date: date, config: AttendanceConfig
) -> tuple[float, bool, str]:
    same_day_punches = [dt for dt in punches if dt.date() == base_date]
    next_day_early_punches = [
        dt
        for dt in punches
        if dt.date() == base_date + timedelta(days=1)
        and dt.time() < config.overnight_cutoff
    ]

    if next_day_early_punches:
        return 4.0, True, "次日凌晨打卡，工作日加班按4小时折算"

    if not same_day_punches:
        return 0.0, False, ""

    last_same_day = max(same_day_punches)
    if last_same_day.time() < config.overtime_start_time:
        return 0.0, False, ""
    if last_same_day.time() >= config.workday_overtime_2h_after:
        return 2.0, last_same_day.time() >= config.workday_meal_after, ""
    return 0.0, False, ""


def weekend_overtime(punches: list[datetime], base_date: date) -> tuple[float, bool]:
    if not punches:
        return 0.0, False

    first_dt = min(punches)
    last_dt = max(punches)
    overtime_periods = [
        (time(9, 0), time(12, 0), 3.0),
        (time(14, 0), time(17, 0), 3.0),
        (time(17, 0), time(21, 0), 2.0),
    ]
    overtime_hours = sum(
        hours
        for start_time, end_time, hours in overtime_periods
        if first_dt <= datetime.combine(base_date, start_time)
        and last_dt >= datetime.combine(base_date, end_time)
    )
    overtime_hours = min(overtime_hours, 8.0)
    return overtime_hours, overtime_hours == 8.0


def holiday_overtime(punches: list[datetime], base_date: date) -> tuple[float, bool]:
    if not punches:
        return 0.0, False

    first_dt = min(punches)
    last_dt = max(punches)
    overtime_periods = [
        (time(9, 0), time(13, 0), 0.5),
        (time(13, 0), time(17, 0), 0.5),
    ]
    overtime_days = sum(
        days
        for start_time, end_time, days in overtime_periods
        if first_dt <= datetime.combine(base_date, start_time)
        and last_dt >= datetime.combine(base_date, end_time)
    )
    overtime_days = min(overtime_days, 1.0)
    return overtime_days, overtime_days == 1.0


def is_late_for_workday(
    punches: list[datetime], base_date: date, config: AttendanceConfig
) -> bool:
    same_day_punches = [dt for dt in punches if dt.date() == base_date]
    if not same_day_punches:
        return False
    return min(same_day_punches).time() > config.work_start_time


def has_normal_start_punch(punches: list[time], config: AttendanceConfig) -> bool:
    return any(config.overnight_cutoff <= punch < config.work_start_time for punch in punches)


def has_non_early_punch(punches: list[time], config: AttendanceConfig) -> bool:
    return any(punch >= config.overnight_cutoff for punch in punches)


def previous_day_is_incomplete(punches: list[time], config: AttendanceConfig) -> bool:
    return len(punches) == 1


def assign_punch_base_date(
    source_date: date,
    punch_time: time,
    natural_day_times: dict[date, list[time]],
    config: AttendanceConfig,
) -> date:
    if punch_time >= config.overnight_cutoff:
        return source_date

    today_times = natural_day_times.get(source_date, [])
    if has_normal_start_punch(today_times, config):
        return source_date - timedelta(days=1)

    if has_non_early_punch(today_times, config):
        return source_date

    previous_times = natural_day_times.get(source_date - timedelta(days=1), [])
    if previous_day_is_incomplete(previous_times, config):
        return source_date - timedelta(days=1)
    return source_date


def style_sheet(ws: Any, freeze_cell: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(45, len(value)))
        ws.column_dimensions[col_letter].width = max(10, min(45, max_len + 2))


def generate_summary(
    input_file: Path,
    output_dir: Path,
    config: AttendanceConfig | None = None,
    *,
    monthly_file: Path | None = None,
) -> GenerationResult:
    active_config = config or AttendanceConfig()
    source = load_workbook(input_file, data_only=True)
    ws = source.active
    report_start, punch_report_end = parse_report_range(ws)
    expected_date_count = (punch_report_end - report_start).days + 1
    if ws.max_column - 6 < expected_date_count:
        source.close()
        raise ValueError("打卡时间表的每日打卡列少于统计日期天数")
    monthly: MonthlyWorkbook | None = None
    if monthly_file is not None:
        monthly = parse_monthly_workbook(monthly_file)
        if (report_start, punch_report_end) != (
            monthly.report_start,
            monthly.report_end,
        ):
            source.close()
            raise ValueError("两张钉钉导出表的统计日期范围不一致")

    source_date_columns: list[tuple[int, date, object]] = []
    date_headers: dict[date, object] = {}
    for offset, col in enumerate(range(7, 7 + expected_date_count)):
        current_date = report_start + timedelta(days=offset)
        header = ws.cell(row=4, column=col).value
        source_date_columns.append((col, current_date, header))
        date_headers[current_date] = header

    ignored_dates = set(active_config.ignore_dates)
    display_dates: list[date] = []
    for offset in range(expected_date_count):
        current = report_start + timedelta(days=offset)
        if active_config.start_date is not None and current < active_config.start_date:
            continue
        if active_config.end_date is not None and current > active_config.end_date:
            continue
        display_dates.append(current)
    report_dates = [current for current in display_dates if current not in ignored_dates]

    detail_rows: list[DetailRow] = []
    person_summary: dict[tuple[str, str, str], dict[str, Any]] = {}
    punch_key_by_user_id: dict[str, tuple[str, str, str]] = {}

    for row in range(5, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        department = str(ws.cell(row=row, column=3).value or "").strip()
        employee_id = str(ws.cell(row=row, column=4).value or "").strip()
        user_id = str(ws.cell(row=row, column=6).value or "").strip()
        key = (name, department, employee_id)
        if user_id:
            punch_key_by_user_id[user_id] = key
        person_summary.setdefault(
            key,
            {
                "ot_dates": [],
                "absence_dates": [],
                "late_dates": [],
                "meal_dates": [],
                "workday_hours": 0.0,
                "weekend_hours": 0.0,
                "holiday_days": 0.0,
                "total_hours": 0.0,
                "meal_count": 0,
            },
        )

        natural_day_times: dict[date, list[time]] = defaultdict(list)
        raw_cells: list[tuple[date, list[tuple[date, time]]]] = []
        for col, source_date, _header in source_date_columns:
            raw_value = ws.cell(row=row, column=col).value
            times = extract_times(raw_value)
            if not times:
                continue
            expanded_times = expand_cell_timeline(source_date, times, active_config)
            for actual_date, punch_time in expanded_times:
                natural_day_times[actual_date].append(punch_time)
            raw_cells.append((source_date, expanded_times))

        assigned: dict[date, dict[str, list[tuple[datetime, str]]]] = defaultdict(
            lambda: {"punches": []}
        )
        for _source_date, expanded_times in raw_cells:
            for actual_date, punch_time in expanded_times:
                base_date = assign_punch_base_date(
                    actual_date, punch_time, natural_day_times, active_config
                )
                punch_dt = datetime.combine(actual_date, punch_time)
                if base_date in report_dates:
                    assigned[base_date]["punches"].append(
                        (punch_dt, f"{fmt_date(actual_date)} {punch_time:%H:%M}")
                    )

        for current_date in report_dates:
            day_data = assigned.get(current_date)
            if not day_data:
                continue
            punch_items = sorted(day_data["punches"], key=lambda item: item[0])
            punches = [item[0] for item in punch_items]
            raw_parts = [item[1] for item in punch_items]
            header = date_headers.get(current_date)
            day_type = get_day_type(current_date, header)
            rest_day = day_type != "工作日"
            absence = False
            overtime_hours = 0.0
            holiday_overtime_days = 0.0
            meal = False
            note = ""

            last_dt = punches[-1]
            last_punch = fmt_dt_for_shift(last_dt, current_date)
            late = False if rest_day else is_late_for_workday(punches, current_date, active_config)

            if len(punches) == 1 and not rest_day:
                absence = True
                note = "工作日只有一次打卡，按旷工标记"
            else:
                if day_type in WEEKEND_DAY_TYPES:
                    overtime_hours, meal = weekend_overtime(punches, current_date)
                elif rest_day:
                    holiday_overtime_days, meal = holiday_overtime(punches, current_date)
                else:
                    overtime_hours, meal, note = workday_overtime(
                        punches, current_date, active_config
                    )
                if last_dt.date() > current_date:
                    note = note or "次日凌晨打卡，按跨天记录处理"

            if overtime_hours > 0 or holiday_overtime_days > 0 or meal or absence or late:
                detail_rows.append(
                    DetailRow(
                        name=name,
                        department=department,
                        employee_id=employee_id,
                        work_date=current_date,
                        day_type=day_type,
                        raw_punches="\n".join(raw_parts),
                        last_punch=last_punch,
                        overtime_hours=overtime_hours,
                        holiday_overtime_days=holiday_overtime_days,
                        meal=meal,
                        meal_amount=active_config.meal_allowance_amount if meal else 0,
                        absent=absence,
                        late=late,
                        note=note,
                    )
                )

            summary = person_summary[key]
            if overtime_hours > 0:
                summary["ot_dates"].append(f"{fmt_date(current_date)}加班{overtime_hours:g}小时")
                if day_type == "工作日":
                    summary["workday_hours"] = float(summary["workday_hours"]) + overtime_hours
                else:
                    summary["weekend_hours"] = float(summary["weekend_hours"]) + overtime_hours
                summary["total_hours"] = float(summary["total_hours"]) + overtime_hours
            if holiday_overtime_days > 0:
                summary["ot_dates"].append(
                    f"{fmt_date(current_date)}加班{holiday_overtime_days:g}天"
                )
                summary["holiday_days"] = (
                    float(summary["holiday_days"]) + holiday_overtime_days
                )
            if absence:
                summary["absence_dates"].append(f"{fmt_date(current_date)}旷工")
            if late:
                summary["late_dates"].append(f"{fmt_date(current_date)}迟到")
            if meal:
                summary["meal_dates"].append(f"{fmt_date(current_date)}有餐补")
                summary["meal_count"] = int(summary["meal_count"]) + 1

    source.close()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / active_config.output_filename
    computed_by_user_id: dict[str, dict[str, Any]] = {}
    for user_id, key in punch_key_by_user_id.items():
        summary = person_summary[key]
        computed_by_user_id[user_id] = {
            "workday_hours": summary["workday_hours"],
            "weekend_hours": summary["weekend_hours"],
            "holiday_days": summary["holiday_days"],
            "meal_count": summary["meal_count"],
            "meal_amount": int(summary["meal_count"])
            * active_config.meal_allowance_amount,
        }
    write_workbook(
        output_path,
        detail_rows,
        person_summary,
        active_config,
        monthly=monthly,
        display_dates=display_dates,
        date_headers=date_headers,
        computed_by_user_id=computed_by_user_id,
        ignored_dates=ignored_dates,
    )
    stats = build_stats(detail_rows, len(person_summary))
    return GenerationResult(output_path=output_path, detail_rows=detail_rows, stats=stats)


def build_stats(detail_rows: list[DetailRow], people: int) -> SummaryStats:
    meal_records = sum(1 for row in detail_rows if row.meal)
    return SummaryStats(
        people=people,
        detail_records=len(detail_rows),
        overtime_records=sum(
            1
            for row in detail_rows
            if row.overtime_hours > 0 or row.holiday_overtime_days > 0
        ),
        absence_records=sum(1 for row in detail_rows if row.absent),
        late_records=sum(1 for row in detail_rows if row.late),
        meal_records=meal_records,
        total_overtime_hours=round(sum(row.overtime_hours for row in detail_rows), 2),
        total_holiday_overtime_days=round(
            sum(row.holiday_overtime_days for row in detail_rows), 2
        ),
        total_meal_amount=sum(row.meal_amount for row in detail_rows),
    )


def write_workbook(
    output_path: Path,
    detail_rows: list[DetailRow],
    person_summary: dict[tuple[str, str, str], dict[str, Any]],
    config: AttendanceConfig,
    *,
    monthly: MonthlyWorkbook | None = None,
    display_dates: list[date] | None = None,
    date_headers: dict[date, object] | None = None,
    computed_by_user_id: dict[str, dict[str, Any]] | None = None,
    ignored_dates: set[date] | None = None,
) -> None:
    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = "加班明细"
    detail_ws.append(
        [
            "姓名",
            "部门",
            "工号",
            "日期",
            "日期类型",
            "原始打卡时间",
            "下班最后打卡时间",
            "折算加班时长",
            "折算节假日加班天数",
            "是否餐补",
            "餐补金额",
            "是否旷工",
            "是否迟到",
            "备注",
        ]
    )
    for row in detail_rows:
        detail_ws.append(
            [
                row.name,
                row.department,
                row.employee_id,
                row.work_date,
                row.day_type,
                row.raw_punches,
                row.last_punch,
                row.overtime_hours,
                row.holiday_overtime_days,
                "是" if row.meal else "否",
                row.meal_amount,
                "是" if row.absent else "否",
                "是" if row.late else "否",
                row.note,
            ]
        )
    for cell in detail_ws["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in detail_ws["H"][1:]:
        cell.number_format = "0.00"
    for cell in detail_ws["I"][1:]:
        cell.number_format = "0.00"
    for cell in detail_ws["K"][1:]:
        cell.number_format = "#,##0"
    style_sheet(detail_ws, "A2")

    summary_ws = wb.create_sheet("人员汇总")
    summary_ws.append(
        [
            "姓名",
            "部门",
            "工号",
            "加班日期",
            "旷工日期",
            "迟到日期",
            "餐补日期",
            "工作日加班时长",
            "周末加班时长",
            "节假日加班天数",
            "工作日及周末加班时长合计",
            "餐补次数",
            "餐补金额合计",
        ]
    )
    for (name, department, employee_id), summary in sorted(person_summary.items()):
        workday_hours = round(float(summary["workday_hours"]), 2)
        weekend_hours = round(float(summary["weekend_hours"]), 2)
        holiday_days = round(float(summary["holiday_days"]), 2)
        total_hours = round(float(summary["total_hours"]), 2)
        meal_count = int(summary["meal_count"])
        if (
            total_hours == 0
            and holiday_days == 0
            and meal_count == 0
            and not summary["absence_dates"]
            and not summary["late_dates"]
        ):
            continue
        summary_ws.append(
            [
                name,
                department,
                employee_id,
                "\n".join(summary["ot_dates"]),
                "\n".join(summary["absence_dates"]),
                "\n".join(summary["late_dates"]),
                "\n".join(summary["meal_dates"]),
                workday_hours,
                weekend_hours,
                holiday_days,
                total_hours,
                meal_count,
                meal_count * config.meal_allowance_amount,
            ]
        )
    for column in ("H", "I", "J", "K"):
        for cell in summary_ws[column][1:]:
            cell.number_format = "0.00"
    for cell in summary_ws["L"][1:]:
        cell.number_format = "#,##0"
    for cell in summary_ws["M"][1:]:
        cell.number_format = "#,##0"
    style_sheet(summary_ws, "A2")

    rules_ws = wb.create_sheet("统计口径")
    rules = [
        ["项目", "口径"],
        ["统计范围", "默认忽略配置中的日期；样例默认忽略2026年6月29日"],
        ["工作日", "表头为空或为阿拉伯数字时按工作日处理，用于支持周末补班"],
        ["周末", "表头为“六”“日”“周六”“周日”“周末”时按周末处理，不按真实星期几推断"],
        ["节假日", "表头为非数字、且不是周末标记的文本时按节假日处理，如“xx节”“法假”"],
        ["工作日加班", "出勤至21:00且不到次日00:00按2小时折算；次日凌晨打卡截止前有有效跨天打卡按4小时折算"],
        ["工作日餐补", "出勤至22:00或更晚享受餐补；有效跨天至次日的4小时加班也享受餐补"],
        ["周末加班", "完整覆盖09:00-12:00算3小时，完整覆盖14:00-17:00算3小时，完整覆盖17:00-21:00算2小时；各时段可累加，每天最多8小时"],
        ["周末餐补", "加班折算满8小时，享受餐补"],
        ["节假日加班", "完整覆盖09:00-13:00算0.5天，完整覆盖13:00-17:00算0.5天；两个时段可累加，每天最多1天"],
        ["节假日餐补", "加班折算满1天，享受餐补"],
        ["迟到", "工作日当天第一条凌晨打卡截止及之后的打卡晚于上班时间，标记迟到；周末/节假日不判迟到"],
        ["凌晨打卡截止", "次日凌晨打卡截止前打卡是跨天加班的打卡约束"],
        ["旷工", "工作日只有一个打卡时间视为旷工；周末/节假日只有一个打卡时间不视为旷工"],
    ]
    for rule_row in rules:
        rules_ws.append(rule_row)
    style_sheet(rules_ws, "A2")

    if monthly is not None:
        if not display_dates:
            raise ValueError("统计日期范围内没有可输出的日期")
        write_attendance_summary_sheet(
            wb,
            monthly,
            display_dates,
            date_headers or {},
            computed_by_user_id or {},
            ignored_dates or set(),
        )

    wb.save(output_path)
