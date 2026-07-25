from __future__ import annotations

import io
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from attendance_sidecar.main import response_for, serve


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "打卡时间.xlsx"


def create_monthly_workbook(path: Path) -> Path:
    punch = load_workbook(SOURCE, data_only=True, read_only=True).active
    wb = Workbook()
    ws = wb.active
    ws.title = "月度汇总"
    ws["A1"] = "月度汇总 2026-06-01 至 2026-06-29"
    ws.cell(row=3, column=37, value="考勤结果")
    for day in range(1, 30):
        ws.cell(row=4, column=36 + day, value=day)
    for target_row, source_row in enumerate(range(5, punch.max_row + 1), start=5):
        if not punch.cell(row=source_row, column=1).value:
            continue
        for col in (1, 3, 4, 6):
            ws.cell(row=target_row, column=col, value=punch.cell(row=source_row, column=col).value)
    wb.save(path)
    wb.close()
    return path


def test_worker_hello() -> None:
    response = response_for(
        {"request_id": "hello-1", "method": "hello", "payload": {}}
    )

    assert response == {
        "request_id": "hello-1",
        "ok": True,
        "result": {"protocol_version": 1, "worker_version": "0.1.0"},
    }


def test_worker_parse_and_generate(tmp_path: Path) -> None:
    monthly_path = create_monthly_workbook(tmp_path / "月度汇总.xlsx")
    parsed = response_for(
        {
            "request_id": "parse-1",
            "method": "parse",
            "payload": {
                "input_path": str(SOURCE),
                "monthly_path": str(monthly_path),
            },
        }
    )
    output_path = tmp_path / "桌面端汇总.xlsx"
    generated = response_for(
        {
            "request_id": "generate-1",
            "method": "generate",
            "payload": {
                "input_path": str(SOURCE),
                "monthly_path": str(monthly_path),
                "output_path": str(output_path),
                "config": {
                    "start_date": "2026-06-01",
                    "end_date": "2026-06-28",
                    "ignore_dates": ["2026-06-29"],
                    "overnight_cutoff": "07:00:00",
                    "work_start_time": "08:30:00",
                    "work_end_time": "18:00:00",
                    "overtime_start_time": "19:00:00",
                    "workday_overtime_2h_after": "21:00:00",
                    "workday_meal_after": "22:00:00",
                    "meal_allowance_amount": 30,
                    "output_filename": "ignored.xlsx",
                },
            },
        }
    )

    assert parsed["ok"] is True
    assert parsed["result"]["filename"] == SOURCE.name
    assert parsed["result"]["monthly_filename"] == monthly_path.name
    assert parsed["result"]["detected"]["employee_count"] == 75
    assert parsed["result"]["detected"]["matched_employee_count"] == 75
    assert generated["ok"] is True
    assert generated["result"]["output_path"] == str(output_path)
    assert generated["result"]["stats"]["late_records"] == 197
    assert output_path.read_bytes().startswith(b"PK")


def test_worker_rejects_overwriting_source() -> None:
    response = response_for(
        {
            "request_id": "generate-1",
            "method": "generate",
            "payload": {
                "input_path": str(SOURCE),
                "monthly_path": str(SOURCE),
                "output_path": str(SOURCE),
                "config": {},
            },
        }
    )

    assert response["ok"] is False
    assert response["error"]["code"] == "invalid_output"


def test_worker_stream_protocol_and_shutdown() -> None:
    source = io.StringIO(
        "not-json\n"
        + json.dumps({"request_id": "bye", "method": "shutdown"})
        + "\n"
    )
    target = io.StringIO()

    serve(source, target)

    responses = [json.loads(line) for line in target.getvalue().splitlines()]
    assert responses[0]["error"]["code"] == "invalid_json"
    assert responses[1] == {
        "request_id": "bye",
        "ok": True,
        "result": {"shutting_down": True},
    }
