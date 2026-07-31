from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from attendance_core.monthly import (
    MonthlyEmployee,
    derive_half_day_statuses,
    employee_leave_days,
    parse_report_range_text,
)
from attendance_core.processor import (
    assign_punch_base_date,
    expand_cell_timeline,
    extract_times,
    generate_summary,
    get_day_type,
    holiday_overtime,
    is_late_for_workday,
    is_holiday_header,
    parse_workbook,
    parse_sources,
    previous_day_is_incomplete,
    weekend_overtime,
    workday_overtime,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "打卡时间.xlsx"


def create_dual_source_workbooks(directory: Path) -> tuple[Path, Path]:
    punch_path = directory / "公司_打卡时间_20260601-20260603.xlsx"
    monthly_path = directory / "公司_月度汇总_20260601-20260603.xlsx"

    punch = Workbook()
    punch_ws = punch.active
    punch_ws["A1"] = "打卡时间 2026-06-01 至 2026-06-03"
    for col, value in enumerate(("1", "六", "端午节"), start=7):
        punch_ws.cell(row=4, column=col, value=value)
    for col, value in {1: "测试员工", 3: "测试部门", 4: "001", 6: "U001"}.items():
        punch_ws.cell(row=5, column=col, value=value)
    punch_ws.cell(row=5, column=7, value="08:20 22:00")
    punch_ws.cell(row=5, column=8, value="09:00 21:00")
    punch_ws.cell(row=5, column=9, value="09:00 17:00")
    punch.save(punch_path)
    punch.close()

    monthly = Workbook()
    monthly_ws = monthly.active
    monthly_ws.title = "月度汇总"
    monthly_ws["A1"] = "月度汇总 2026-06-01 至 2026-06-03"
    monthly_ws.cell(row=3, column=34, value="旷工天数")
    monthly_ws.cell(row=4, column=22, value="事假(小时)")
    monthly_ws.cell(row=4, column=25, value="年假(天)")
    monthly_ws.cell(row=3, column=37, value="考勤结果")
    for col, value in enumerate(("1", "2", "3"), start=37):
        monthly_ws.cell(row=4, column=col, value=value)
    for col, value in {1: "测试员工", 3: "测试部门", 4: "001", 6: "U001"}.items():
        monthly_ws.cell(row=5, column=col, value=value)
    monthly_ws.cell(row=5, column=37, value="正常\n(08:20,22:00)")
    monthly_ws.cell(row=5, column=38, value="休息\n(09:00,21:00)")
    monthly_ws.cell(row=5, column=39, value="休息\n(09:00,17:00)")
    monthly.save(monthly_path)
    monthly.close()
    return punch_path, monthly_path


def test_monthly_half_day_status_uses_official_period_overlap() -> None:
    current = date(2026, 6, 15)
    assert derive_half_day_statuses(
        "年假06-15 08:30到06-15 13:15 0.5天\n(-,18:32)", current, "工作日"
    ) == ("年假", "√")
    assert derive_half_day_statuses(
        "事假06-15 13:15到06-15 18:00 0.5天\n(08:23,12:01)", current, "工作日"
    ) == ("√", "事假")
    assert derive_half_day_statuses("出差\n(-,-)", current, "工作日") == ("√", "√")
    assert derive_half_day_statuses("外出\n(-,-)", current, "工作日") == ("√", "√")
    assert derive_half_day_statuses("外勤\n(-,-)", current, "工作日") == ("√", "√")


def test_hourly_leave_totals_are_counted_from_half_day_intervals() -> None:
    display_dates = [
        date(2026, 7, 10) + timedelta(days=offset)
        for offset in range(8)
    ]
    date_headers = {
        current_date: (
            "六"
            if current_date.weekday() == 5
            else "日"
            if current_date.weekday() == 6
            else str(current_date.day)
        )
        for current_date in display_dates
    }
    employee = MonthlyEmployee(
        name="测试员工",
        department="测试部门",
        employee_id="001",
        user_id="U001",
        leave_totals={
            "事假": 51,
            "调休": 4,
            "病假": 4,
            "哺乳假": 4,
            "年假": 1,
        },
        daily_results={
            current_date: (
                "事假07-10 09:00到07-15 18:00 4天\n(-)"
                if current_date <= date(2026, 7, 15)
                else "事假07-16 08:30到07-17 18:00 2天\n(-)"
            )
            for current_date in display_dates
        },
    )

    leave_days = employee_leave_days(employee, display_dates, date_headers)

    assert leave_days["事假"] == 6
    assert leave_days["调休"] == 0
    assert leave_days["病假"] == 0
    assert leave_days["哺乳假"] == 0
    assert leave_days["年假"] == 1


def test_hourly_leave_counts_each_matching_period_as_half_day() -> None:
    morning_date = date(2026, 7, 20)
    afternoon_date = date(2026, 7, 21)
    inactive_date = date(2026, 7, 22)
    display_dates = [morning_date, afternoon_date, inactive_date]
    employee = MonthlyEmployee(
        name="测试员工",
        department="测试部门",
        employee_id="001",
        user_id="U001",
        leave_totals={"事假": 12, "病假": 4},
        daily_results={
            morning_date: "事假07-20 08:30到07-20 13:15 0.5天\n(-,18:00)",
            afternoon_date: "病假07-21 13:15到07-21 18:00 0.5天\n(08:20,-)",
            inactive_date: "事假07-22 08:30到07-22 18:00 1天\n(-)",
        },
    )

    leave_days = employee_leave_days(
        employee,
        display_dates,
        {current_date: str(current_date.day) for current_date in display_dates},
        {inactive_date: "已离职"},
    )

    assert leave_days["事假"] == 0.5
    assert leave_days["病假"] == 0.5


def test_monthly_half_day_status_marks_non_workdays() -> None:
    current = date(2026, 6, 14)
    punches = [datetime.combine(current, time(9, 0))]

    for day_type in ("周六", "周日", "周末", "端午节"):
        assert derive_half_day_statuses("休息\n(-,-)", current, day_type) == (
            "○",
            "○",
        )
        assert derive_half_day_statuses(
            "休息\n(-,-)",
            current,
            day_type,
            [],
        ) == (
            "○",
            "○",
        )
        assert derive_half_day_statuses(
            "休息\n(-,-)",
            current,
            day_type,
            punches,
        ) == (
            "√",
            "√",
        )


def test_generated_summary_marks_non_workdays_without_punches_as_circles(
    tmp_path: Path,
) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch_ws = punch.active
    punch_ws.cell(row=5, column=8).value = None
    punch_ws.cell(row=5, column=9).value = None
    punch.save(punch_path)
    punch.close()

    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="非工作日无打卡.xlsx",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    summary = workbook["汇总表"]

    assert summary["O5"].value == "√"
    assert summary["O6"].value == "√"
    for coordinate in ("P5", "P6", "Q5", "Q6"):
        assert summary[coordinate].value == "○"
    workbook.close()


def test_half_day_status_prefers_actual_punches_over_monthly_label() -> None:
    current = date(2026, 7, 1)

    assert derive_half_day_statuses(
        "出差06-16 08:30到07-23 13:15 37.5天\n(-)",
        current,
        "工作日",
        [],
    ) == ("异常", "异常")
    assert derive_half_day_statuses(
        "正常\n(08:30,18:00)",
        current,
        "工作日",
        [
            datetime.combine(current, time(8, 30)),
            datetime.combine(current, time(18, 0)),
        ],
    ) == ("√", "√")


def test_dual_sources_generate_four_sheet_summary(tmp_path: Path) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    parsed = parse_sources(punch_path, monthly_path)
    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="考勤汇总.xlsx",
    )

    assert parsed.matched_employee_count == 1
    assert not parsed.source_warnings
    workbook = load_workbook(result.output_path, data_only=True)
    assert workbook.sheetnames == ["汇总表", "加班明细", "人员汇总", "统计口径"]
    summary = workbook["汇总表"]
    assert summary["C5"].value == "测试员工"
    assert summary["D5"].value == 1
    assert summary["E5"].value == 2
    assert summary["F5"].value == 8
    assert summary["O5"].value == "√"
    assert summary["O6"].value == "√"
    assert summary["P5"].value == "√"
    assert summary["P6"].value == "√"
    assert summary["Q5"].value == "√"
    assert summary["Q6"].value == "√"
    workbook.close()


def test_generated_summary_uses_punch_sheet_for_daily_status(tmp_path: Path) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch.active.cell(row=5, column=7).value = None
    punch.save(punch_path)
    punch.close()

    monthly = load_workbook(monthly_path)
    monthly.active.cell(
        row=5,
        column=37,
        value="出差05-20 08:30到06-02 18:00 14天\n(-)",
    )
    monthly.save(monthly_path)
    monthly.close()

    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="打卡优先.xlsx",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    summary = workbook["汇总表"]

    assert summary["O5"].value == "异常"
    assert summary["O6"].value == "异常"
    workbook.close()


def test_employment_boundary_markers_expand_into_daily_statuses(
    tmp_path: Path,
) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch_ws = punch.active
    punch_ws.cell(row=5, column=7, value="未入职")
    punch_ws.cell(row=5, column=9, value="已离职")
    punch.save(punch_path)
    punch.close()

    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="在职边界.xlsx",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    summary = workbook["汇总表"]

    assert summary["O5"].value == "未入职"
    assert summary["O6"].value == "未入职"
    assert summary["P5"].value == "√"
    assert summary["P6"].value == "√"
    assert summary["Q5"].value == "已离职"
    assert summary["Q6"].value == "已离职"
    assert summary.cell(row=5, column=summary.max_column - 1).value == (
        "6/1未入职；6/3已离职；餐补1次/30元"
    )
    assert result.stats.total_overtime_hours == 8
    assert result.stats.total_holiday_overtime_days == 0
    assert result.stats.meal_records == 1
    workbook.close()


def test_last_day_not_hired_marker_covers_entire_report(tmp_path: Path) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch_ws = punch.active
    punch_ws.cell(row=5, column=7).value = None
    punch_ws.cell(row=5, column=8).value = None
    punch_ws.cell(row=5, column=9, value="未入职")
    punch.save(punch_path)
    punch.close()

    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="整月未入职.xlsx",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    summary = workbook["汇总表"]

    for coordinate in ("O5", "O6", "P5", "P6", "Q5", "Q6"):
        assert summary[coordinate].value == "未入职"
    assert summary.cell(row=5, column=summary.max_column - 1).value == (
        "6/1–6/3未入职"
    )
    assert result.stats.overtime_records == 0
    assert result.stats.absence_records == 0
    assert result.stats.late_records == 0
    assert result.stats.meal_records == 0
    workbook.close()


def test_employment_marker_rejects_punches_inside_inactive_period(
    tmp_path: Path,
) -> None:
    punch_path, _monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch.active.cell(row=5, column=8, value="未入职")
    punch.save(punch_path)
    punch.close()

    with pytest.raises(ValueError, match="已标记为“未入职”，但仍存在打卡时间"):
        parse_workbook(punch_path)


def test_employment_markers_require_valid_order(tmp_path: Path) -> None:
    punch_path, _monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch_ws = punch.active
    punch_ws.cell(row=5, column=7, value="已离职")
    punch_ws.cell(row=5, column=8, value="未入职")
    punch_ws.cell(row=5, column=9).value = None
    punch.save(punch_path)
    punch.close()

    with pytest.raises(ValueError, match="“未入职”标记必须早于“已离职”标记"):
        parse_workbook(punch_path)


def test_employment_marker_must_be_the_only_cell_content(tmp_path: Path) -> None:
    punch_path, _monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch.active.cell(row=5, column=7, value="未入职 08:20")
    punch.save(punch_path)
    punch.close()

    with pytest.raises(ValueError, match="在职标记必须单独填写"):
        parse_workbook(punch_path)


def test_single_source_employees_keep_available_information(tmp_path: Path) -> None:
    punch_path, monthly_path = create_dual_source_workbooks(tmp_path)

    punch = load_workbook(punch_path)
    punch_ws = punch.active
    for col, value in {
        1: "仅打卡员工",
        3: "测试部门",
        4: "002",
        6: "U002",
        7: "08:20 22:00",
    }.items():
        punch_ws.cell(row=6, column=col, value=value)
    punch.save(punch_path)
    punch.close()

    monthly = load_workbook(monthly_path)
    monthly_ws = monthly.active
    for col, value in {
        1: "仅月度员工",
        3: "测试部门",
        4: "003",
        6: "U003",
        37: "正常\n(08:20,18:00)",
    }.items():
        monthly_ws.cell(row=6, column=col, value=value)
    monthly.save(monthly_path)
    monthly.close()

    parsed = parse_sources(punch_path, monthly_path)
    result = generate_summary(
        punch_path,
        tmp_path,
        monthly_file=monthly_path,
        output_filename="人员差异.xlsx",
    )
    workbook = load_workbook(result.output_path, data_only=True)

    detail_names = {
        str(cell.value)
        for cell in workbook["加班明细"]["A"][1:]
        if cell.value not in (None, "")
    }
    monthly_names = {
        str(workbook["汇总表"].cell(row=row, column=3).value)
        for row in range(5, workbook["汇总表"].max_row + 1, 2)
    }

    assert parsed.matched_employee_count == 1
    assert parsed.punch_only_employees == ["仅打卡员工"]
    assert parsed.monthly_only_employees == ["仅月度员工"]
    assert "仅打卡员工" in detail_names
    assert "仅打卡员工" in monthly_names
    assert "仅月度员工" in monthly_names
    assert result.stats.people == 3
    punch_only_row = next(
        row
        for row in range(5, workbook["汇总表"].max_row + 1, 2)
        if workbook["汇总表"].cell(row=row, column=3).value == "仅打卡员工"
    )
    assert workbook["汇总表"].cell(row=punch_only_row, column=5).value == 2
    assert workbook["汇总表"].cell(row=punch_only_row, column=15).value is None
    assert (
        workbook["汇总表"].cell(
            row=punch_only_row,
            column=workbook["汇总表"].max_column - 1,
        ).value
        == "餐补1次/30元"
    )
    workbook.close()


def test_sample_workbook_uses_complete_report_range() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    assert any(row.work_date == date(2026, 6, 29) for row in result.detail_rows)


def test_workday_late_boundary_uses_0830() -> None:
    base_date = date(2026, 6, 10)
    assert not is_late_for_workday(
        [datetime.combine(base_date, time(8, 30))], base_date
    )
    assert is_late_for_workday(
        [datetime.combine(base_date, time(8, 31))], base_date
    )


def test_workday_0830_minute_is_normal_through_workbook_processing(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "08点30分边界.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "打卡时间 2026-06-10 至 2026-06-10"
    sheet.cell(row=4, column=7, value="10")
    for col, value in {
        1: "边界员工",
        3: "测试部门",
        4: "001",
        6: "U001",
        7: "08:30:59 21:00",
    }.items():
        sheet.cell(row=5, column=col, value=value)
    workbook.save(source_path)
    workbook.close()

    result = generate_summary(source_path, tmp_path, output_filename="边界结果.xlsx")
    boundary_row = next(
        row
        for row in result.detail_rows
        if row.name == "边界员工" and row.work_date == date(2026, 6, 10)
    )

    assert boundary_row.raw_punches.startswith("6月10日 08:30")
    assert not boundary_row.late


def test_workday_overtime_values_are_only_two_or_four() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")
    workday_hours = {
        row.overtime_hours
        for row in result.detail_rows
        if row.day_type == "工作日" and row.overtime_hours > 0
    }

    assert workday_hours == {2.0, 4.0}


def test_workday_overtime_and_meal_boundaries() -> None:
    base_date = date(2026, 6, 10)
    def calculate(last_dt: datetime) -> tuple[float, bool]:
        hours, meal, _note = workday_overtime(
            [datetime.combine(base_date, time(8, 20)), last_dt],
            base_date,
        )
        return hours, meal

    assert calculate(datetime.combine(base_date, time(20, 59))) == (0.0, False)
    assert calculate(datetime.combine(base_date, time(21, 0))) == (2.0, False)
    assert calculate(datetime.combine(base_date, time(21, 59))) == (2.0, False)
    assert calculate(datetime.combine(base_date, time(22, 0))) == (2.0, True)
    assert calculate(datetime.combine(base_date, time(23, 59))) == (2.0, True)
    assert calculate(datetime.combine(base_date + timedelta(days=1), time(0, 0))) == (4.0, True)


def test_weekend_overtime_requires_full_period_coverage() -> None:
    base_date = date(2026, 6, 13)

    def punches(start: time, end: time) -> list[datetime]:
        return [datetime.combine(base_date, start), datetime.combine(base_date, end)]

    assert weekend_overtime(punches(time(10, 0), time(13, 0)), base_date) == (0.0, False)
    assert weekend_overtime(punches(time(10, 0), time(17, 0)), base_date) == (3.0, False)
    assert weekend_overtime(punches(time(8, 30), time(18, 0)), base_date) == (6.0, False)
    assert weekend_overtime(punches(time(9, 0), time(21, 0)), base_date) == (8.0, True)
    assert weekend_overtime(punches(time(8, 0), time(22, 0)), base_date) == (8.0, True)


def test_holiday_overtime_requires_full_period_coverage() -> None:
    base_date = date(2026, 6, 19)

    def punches(start: time, end: time) -> list[datetime]:
        return [datetime.combine(base_date, start), datetime.combine(base_date, end)]

    assert holiday_overtime(punches(time(10, 0), time(12, 0)), base_date) == (0.0, False)
    assert holiday_overtime(punches(time(10, 0), time(17, 0)), base_date) == (0.5, False)
    assert holiday_overtime(punches(time(9, 0), time(12, 0)), base_date) == (0.5, False)
    assert holiday_overtime(punches(time(14, 0), time(17, 0)), base_date) == (0.5, False)
    assert holiday_overtime(punches(time(9, 0), time(17, 0)), base_date) == (1.0, True)
    assert holiday_overtime(punches(time(9, 1), time(17, 0)), base_date) == (0.5, False)
    assert holiday_overtime(punches(time(8, 59), time(17, 0)), base_date) == (1.0, True)
    assert holiday_overtime(punches(time(10, 0), time(19, 0)), base_date) == (0.5, True)
    assert holiday_overtime(punches(time(8, 0), time(22, 0)), base_date) == (1.0, True)


def test_early_morning_punch_stays_today_when_today_has_no_normal_start_punch() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    rows = {
        (row.name, row.work_date): row
        for row in result.detail_rows
        if row.name in {"柴伟", "郭孟良"}
    }

    chai_24 = rows[("柴伟", date(2026, 6, 24))]
    chai_25 = rows[("柴伟", date(2026, 6, 25))]
    guo_11 = rows[("郭孟良", date(2026, 6, 11))]

    assert chai_24.last_punch == "22:44"
    assert chai_24.overtime_hours == 2.0
    assert "6月25日 06:40" in chai_25.raw_punches
    assert not chai_25.late
    assert chai_25.overtime_hours == 2.0

    assert guo_11.last_punch == "21:14"
    assert guo_11.overtime_hours == 2.0
    assert not any(
        row.name == "郭孟良" and row.work_date == date(2026, 6, 12) and row.absent
        for row in result.detail_rows
    )


def test_previous_day_incomplete_means_only_one_punch() -> None:
    assert previous_day_is_incomplete([time(21, 14)])
    assert not previous_day_is_incomplete([time(8, 25), time(21, 14)])
    assert (
        assign_punch_base_date(
            date(2026, 6, 12),
            time(6, 37),
            {date(2026, 6, 11): [time(8, 25), time(21, 14)], date(2026, 6, 12): [time(6, 37)]},
        )
        == date(2026, 6, 12)
    )
    assert (
        assign_punch_base_date(
            date(2026, 6, 12),
            time(6, 37),
            {
                date(2026, 6, 11): [time(8, 20), time(21, 14)],
                date(2026, 6, 12): [time(6, 37), time(8, 30), time(18, 0)],
            },
        )
        == date(2026, 6, 11)
    )


def test_same_cell_time_rollover_counts_as_next_day_overtime() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    row = next(
        item
        for item in result.detail_rows
        if item.name == "付陈良" and item.work_date == date(2026, 6, 26)
    )

    assert row.raw_punches == "6月26日 07:30\n6月27日 01:54"
    assert row.last_punch == "次日 01:54"
    assert row.overtime_hours == 4.0
    assert row.meal
    assert not row.absent
    assert not any(
        item.name == "付陈良" and item.work_date == date(2026, 6, 25) and "01:54" in item.raw_punches
        for item in result.detail_rows
    )


def test_overnight_notes_are_human_readable() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")
    notes = {row.note for row in result.detail_rows if row.note}

    assert "次日凌晨打卡，工作日加班按4小时折算" in notes
    assert "次日凌晨打卡，按跨天记录处理" in notes
    assert not any("约束已满足" in note for note in notes)


def test_same_cell_next_day_only_uses_trailing_early_off_work_punches() -> None:
    source_date = date(2026, 6, 26)

    assert expand_cell_timeline(source_date, [time(7, 30), time(1, 54)]) == [
        (date(2026, 6, 26), time(7, 30)),
        (date(2026, 6, 27), time(1, 54)),
    ]
    assert expand_cell_timeline(source_date, [time(8, 30), time(8, 20), time(18, 0)]) == [
        (date(2026, 6, 26), time(8, 30)),
        (date(2026, 6, 26), time(8, 20)),
        (date(2026, 6, 26), time(18, 0)),
    ]


def test_extract_times_handles_adjacent_times_without_separator() -> None:
    assert extract_times("08:3022:32外勤") == [time(8, 30), time(22, 32)]


def test_weekend_single_punch_is_not_absent() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    assert not any(
        row.name == "余祖应" and row.work_date == date(2026, 6, 13) and row.absent
        for row in result.detail_rows
    )


def test_holiday_header_marks_day_as_rest_day_without_config_parameter() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    assert any(
        row.work_date == date(2026, 6, 19) and row.day_type == "端午节"
        for row in result.detail_rows
    )


def test_weekend_day_type_uses_specific_weekday_name() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")

    assert any(row.work_date == date(2026, 6, 13) and row.day_type == "周六" for row in result.detail_rows)
    assert any(row.work_date == date(2026, 6, 14) and row.day_type == "周日" for row in result.detail_rows)


def test_day_type_comes_from_header_for_makeup_workdays_and_rest_days() -> None:
    assert get_day_type(date(2026, 6, 13), "13") == "工作日"
    assert get_day_type(date(2026, 6, 10), "六") == "周六"
    assert get_day_type(date(2026, 6, 10), "日") == "周日"
    assert get_day_type(date(2026, 6, 10), "周末") == "周末"
    assert get_day_type(date(2026, 6, 10), "周六") == "周六"
    assert get_day_type(date(2026, 6, 10), "周日") == "周日"
    assert get_day_type(date(2026, 6, 10), "法假") == "法假"
    assert not is_holiday_header("周末")
    assert is_holiday_header("法假")


def test_summary_uses_header_not_weekday_for_makeup_workdays_and_rest_days() -> None:
    source = ROOT / "outputs" / "tests" / "makeup-days.xlsx"
    source.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "考勤报表 2026-06-13 至 2026-06-15"
    ws.cell(row=4, column=7, value="13")
    ws.cell(row=4, column=8, value="周末")
    ws.cell(row=4, column=9, value="法假")
    ws.cell(row=5, column=1, value="测试员工")
    ws.cell(row=5, column=3, value="测试部门")
    ws.cell(row=5, column=4, value="001")
    ws.cell(row=5, column=7, value="08:20")
    ws.cell(row=5, column=8, value="08:30 17:00")
    ws.cell(row=5, column=9, value="08:30 21:30")
    wb.save(source)
    wb.close()

    result = generate_summary(source, ROOT / "outputs" / "tests")
    rows = {row.work_date: row for row in result.detail_rows}

    assert rows[date(2026, 6, 13)].day_type == "工作日"
    assert rows[date(2026, 6, 13)].absent
    assert rows[date(2026, 6, 14)].day_type == "周末"
    assert not rows[date(2026, 6, 14)].absent
    assert rows[date(2026, 6, 14)].overtime_hours == 6.0
    assert rows[date(2026, 6, 15)].day_type == "法假"
    assert rows[date(2026, 6, 15)].holiday_overtime_days == 1.0


def test_person_summary_splits_overtime_hours_by_day_type() -> None:
    result = generate_summary(SOURCE, ROOT / "outputs" / "tests")
    wb = load_workbook(result.output_path, data_only=True)
    ws = wb["人员汇总"]

    headers = [cell.value for cell in ws[1]]
    assert headers[7:11] == [
        "工作日加班时长",
        "周末加班时长",
        "节假日加班天数",
        "工作日及周末加班时长合计",
    ]

    expected: dict[tuple[str, str, str], dict[str, float]] = {}
    for row in result.detail_rows:
        key = (row.name, row.department, row.employee_id)
        person = expected.setdefault(
            key,
            {"workday": 0.0, "weekend": 0.0, "holiday": 0.0, "total": 0.0},
        )
        if row.day_type == "工作日" and row.overtime_hours > 0:
            person["workday"] += row.overtime_hours
            person["total"] += row.overtime_hours
        elif row.day_type in {"周六", "周日"} and row.overtime_hours > 0:
            person["weekend"] += row.overtime_hours
            person["total"] += row.overtime_hours
        elif row.holiday_overtime_days > 0:
            person["holiday"] += row.holiday_overtime_days

    actual_rows = {
        (
            str(row[0].value or ""),
            str(row[1].value or ""),
            str(row[2].value or ""),
        ): row
        for row in ws.iter_rows(min_row=2)
    }
    for key, values in expected.items():
        if values["total"] == 0 and values["holiday"] == 0:
            continue
        row = actual_rows[key]
        assert float(row[7].value or 0) == round(values["workday"], 2)
        assert float(row[8].value or 0) == round(values["weekend"], 2)
        assert float(row[9].value or 0) == round(values["holiday"], 2)
        assert float(row[10].value or 0) == round(values["total"], 2)

    wb.close()


def test_parse_workbook_returns_detected_defaults() -> None:
    parsed = parse_workbook(SOURCE)

    assert parsed.report_start == date(2026, 6, 1)
    assert parsed.report_end == date(2026, 6, 29)
    assert [(item.date, item.label) for item in parsed.holidays] == [
        (date(2026, 6, 19), "端午节")
    ]
    assert (date(2026, 6, 13), "周六") in [
        (item.date, item.label) for item in parsed.non_workdays
    ]
    assert (date(2026, 6, 14), "周日") in [
        (item.date, item.label) for item in parsed.non_workdays
    ]
    assert (date(2026, 6, 19), "端午节") in [
        (item.date, item.label) for item in parsed.non_workdays
    ]
    assert parsed.employee_count == 75
    assert parsed.date_count == 29


def test_report_titles_require_complete_ordered_date_ranges(tmp_path: Path) -> None:
    punch_path = tmp_path / "打卡时间.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "打卡时间 2026-06-01"
    workbook.save(punch_path)
    workbook.close()

    with pytest.raises(ValueError, match="打卡时间表标题中未找到完整统计日期范围"):
        parse_workbook(punch_path)

    with pytest.raises(ValueError, match="打卡时间表标题中的统计结束日期早于开始日期"):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "打卡时间 2026-06-30 至 2026-06-01"
        workbook.save(punch_path)
        workbook.close()
        parse_workbook(punch_path)

    with pytest.raises(ValueError, match="月度汇总表标题中未找到完整统计日期范围"):
        parse_report_range_text("月度汇总 2026-06-01")

    with pytest.raises(ValueError, match="月度汇总表标题中的统计结束日期早于开始日期"):
        parse_report_range_text("月度汇总 2026-06-30 至 2026-06-01")
